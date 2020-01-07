from openpyxl import Workbook
from openpyxl.styles import PatternFill
import cv2


def main():
    wb = Workbook()
    ws = wb.active

    img = cv2.imread('22.jpg', cv2.IMREAD_COLOR)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    img_shape = img.shape
    empty = ['' for _ in range(img_shape[1])]

    ws_info_arr = []
    rgb_arr = []

    for i in range(img_shape[0]):
        ws.append(empty)
        
        for j in range(img_shape[1]):  
            color = img[i, j]
            rgb = [hex(c)[2:].upper() for c in color]
            
            for idx, r in enumerate(rgb):
                if len(r) < 2:
                    t = '0' + r
                    rgb[idx] = t
                    
            argb = 'FF' + ''.join(rgb)
            rgb_arr.append(argb)
        
            ws_info_arr.append((i, j, argb))    

    rgb_dict = {x:  PatternFill(start_color=x, end_color=x, fill_type='solid') 
            for x in list(set(rgb_arr))}

    for info in ws_info_arr:
        cell = ws.cell(info[0] + 1, info[1] + 1)
        cell.fill = rgb_dict[info[2]]

    col_names = [s.column_letter for s in ws[1]]    

    for col in col_names:
        ws.column_dimensions[col].width = 1
        
    for row in range(1, img_shape[0] + 1):
        ws.row_dimensions[row].height = 6
            
    wb.save('./xx.xlsx')
    wb.close()

main()
